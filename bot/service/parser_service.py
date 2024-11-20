import re
import logging
from typing import Dict, List, Optional

import openpyxl
from openpyxl.cell import MergedCell, Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from bot.entity.enum.DayOfWeek import DayOfWeek
from bot.entity.enum.WeekType import WeekType
from bot.entity.shedule.TimeInterval import TimeInterval
from bot.entity.shedule.UniversityShedule import UniversitySchedule

# Настройка логгера
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class ParserService:

    def parse_file(self, xlsx_location: str) -> UniversitySchedule:
        logger.info("Начинается парсинг файла: %s", xlsx_location)
        wb: Workbook = openpyxl.load_workbook(filename=xlsx_location)
        us: UniversitySchedule = UniversitySchedule()
        for ws in wb.worksheets:
            logger.info("Парсинг листа: %s", ws.title)
            us.merge_schedules(self.__parse_worksheet(ws))
        return us

    def __parse_worksheet(self, ws: Worksheet):
        logger.debug("Ищем ячейку с текстом 'дни'")
        left_up_col, left_up_row = -1, -1
        us: UniversitySchedule = UniversitySchedule()
        for j in range(1, 11):
            for i in range(1, 11):
                cell = ws.cell(row=i, column=j)
                if cell is not None and cell.value is not None and cell.value.strip().lower() == 'дни':
                    left_up_row, left_up_col = i, j
                    logger.info("Найдена ячейка: колонка %d, строка %d", left_up_col, left_up_row)
                    break
        if left_up_col == -1 or left_up_row == -1:
            logger.warning("Не удалось найти ячейку с текстом 'дни'")
            return us

        # Парсим группы и номер колонки для каждой
        DAY_COLUMN = left_up_col
        TIME_COLUMN = DAY_COLUMN + 1
        WEEK_TYPE_COLUMN = DAY_COLUMN + 2
        column_groups: Dict[str, int] = {}
        for col in range(left_up_col + 1, ws.max_column + 1):
            cell = ws.cell(column=col, row=left_up_row)
            if cell is not None and cell.value is not None:
                value = str(cell.value).strip().lower()
                digit_count = 0
                for char in value:
                    if char.isdigit():
                        digit_count += 1
                        if digit_count >= 2:
                            group = cell.value.strip().lower()
                            sm_ind = group.lower().find('см')
                            if sm_ind != -1:
                                group = group[:sm_ind]
                            column_groups[group] = cell.column
                            break
        logger.info("Обнаружены группы: %s", ", ".join(column_groups.keys()))
        for group in column_groups:
            us.add_group(group)

        time, day = None, None
        prev_subject: Dict[str, Optional[None, str]] = {group: None for group in column_groups}
        prev_cell: Dict[str, Optional[None, Cell]] = {group: None for group in column_groups}

        # Парсим расписания
        for row in range(left_up_row + 1, ws.max_row + 1):
            week_type_str = ws.cell(row, WEEK_TYPE_COLUMN).value
            if week_type_str is None:
                logger.info("Достигнут конец таблицы расписания на строке %d", row - 1)
                break
            if 'числ' in week_type_str.lower():
                week_type = WeekType.ODD_WEEK
            elif 'знам' in week_type_str.lower():
                week_type = WeekType.EVEN_WEEK
            else:
                logger.error("Неверный формат типа недели. Строка: %d, столбец: %d", row, DAY_COLUMN)
                raise Exception(f"Неверный формат типа недели. Строка: {row}, столбец {DAY_COLUMN}")

            if ws.cell(row, DAY_COLUMN).value is not None:
                prev_subject = {group: None for group in column_groups}
                day = DayOfWeek.get_day_from_string(ws.cell(row, DAY_COLUMN).value)
                logger.debug("Обнаружен день недели: %s", day)

            if ws.cell(row, TIME_COLUMN).value is not None:
                time = TimeInterval.from_dirty_string(ws.cell(row, TIME_COLUMN).value)
                logger.debug("Обнаружено время: %s", time)

            for group, column in column_groups.items():
                curr_cell = ws.cell(row, column)
                subject: str = curr_cell.value
                if prev_cell[group] is not None and are_cells_merged(prev_cell[group], curr_cell,ws):
                    subject = prev_subject[group]
                else:
                    prev_subject[group] = subject
                if subject is None or len(subject.strip()) < 2:
                    continue
                prev_cell[group] = curr_cell
                us.get_group_day_schedule(group, day, week_type).add_lesson(time, subject)
        logger.info("Расписание для листа %s успешно спарсено", ws.title)
        return us


def are_cells_merged(cell1: Cell, cell2: Cell, sheet: Worksheet):
    # Проверяем все диапазоны объединенных ячеек
    for merged_range in sheet.merged_cells.ranges:
        # Если обе ячейки находятся в одном диапазоне объединенных ячеек
        if cell1.coordinate in merged_range and cell2.coordinate in merged_range:
            return True
    return False
