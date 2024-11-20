import re
from typing import Dict, List

import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from bot.entity.enum.DayOfWeek import DayOfWeek
from bot.entity.enum.WeekType import WeekType
from bot.entity.shedule.TimeInterval import TimeInterval
from bot.entity.shedule.UniversityShedule import UniversitySchedule
from pprint import pprint


class Parser:

    def parse_file(self, xlsx_location: str) -> UniversitySchedule:
        wb: Workbook = openpyxl.load_workbook(filename=xlsx_location)
        us: UniversitySchedule = UniversitySchedule()
        for ws in wb.worksheets:
            us.merge_schedules(self.__parse_worksheet(ws))
        return us

    def __parse_worksheet(self, ws: Worksheet):
        # Ищем ячейку, в которой написано 'дни'
        left_up_col, left_up_row = -1, -1
        us: UniversitySchedule = UniversitySchedule()
        for j in range(1, 11):
            for i in range(1, 11):
                cell = ws.cell(row=i, column=j)
                if cell is not None and \
                        cell.value is not None and \
                        cell.value.strip().lower() == 'дни':
                    left_up_row, left_up_col = i, j
                    print('колонка ', left_up_col, 'строка ', left_up_row)
                    break
        if left_up_col == -1 or left_up_row == -1:
            return us

        # Парсим группы и номер колонки для каждой
        DAY_COLUMN = left_up_col
        TIME_COLUMN = DAY_COLUMN + 1
        WEEK_TYPE_COLUMN = DAY_COLUMN + 2
        column_groups: Dict[str, int] = {}  # Номер группы->колонка
        for col in range(left_up_col + 1, ws.max_column + 1):
            cell = ws.cell(column=col, row=left_up_row)
            # print(cell.row,cell.column)
            if cell is not None and cell.value is not None:
                value = str(cell.value).strip().lower()  # Преобразуем значение в строку, чтобы работать с ним
                digit_count = 0  # Считаем количество цифр
                # Проходим по символам строки
                for char in value:
                    if char.isdigit():
                        digit_count += 1
                        # Если цифр стало две, можно прервать цикл
                        if digit_count >= 2:
                            group = cell.value.strip().lower()
                            sm_ind = group.lower().find('см')
                            if sm_ind != -1:
                                group = group[:sm_ind]
                            column_groups[group] = cell.column
                            break  # Прерываем проверку после нахождения двух цифр
        # print(column_groups)
        for group in column_groups:
            us.add_group(group)

        time, day = None, None
        col_prev_subject = {group: None for group in column_groups}
        # Парсим расписания
        prev_merged = False
        for row in range(left_up_row + 1, ws.max_row + 1):
            # Определяем тип недели и проверяем  конец таблицы расписания
            week_type_str = ws.cell(row, WEEK_TYPE_COLUMN).value
            if week_type_str is None:
                # print('Последняя строка расписания', row - 1)
                break
            if 'числ' in week_type_str.lower():
                week_type = WeekType.ODD_WEEK
            elif 'знам' in week_type_str.lower():
                week_type = WeekType.EVEN_WEEK
            else:
                raise Exception(f'Неверный формат типа недели.строка: {row}, столбец {DAY_COLUMN}')

            # Определим день недели
            if ws.cell(row, DAY_COLUMN).value is not None:
                col_prev_subject = {group: None for group in column_groups}
                day = DayOfWeek.get_day_from_string(ws.cell(row, DAY_COLUMN).value)
                # print(day, 'колонка ', DAY_COLUMN, 'строка ', row)

            # Определяем время для пары
            if ws.cell(row, TIME_COLUMN).value is not None:
                time = TimeInterval.from_dirty_string(ws.cell(row, TIME_COLUMN).value)
                # print(time)

            # Расписание для каждой группы
            for group, column in column_groups.items():
                prev_subject = col_prev_subject[group]
                subject_cell = ws.cell(row, column)
                subject = subject = subject_cell.value
                if isinstance(subject_cell, MergedCell):
                    print(subject_cell.value)
                    if not prev_merged:
                        prev_subject = None
                    else:
                        subject = prev_subject
                    prev_merged = True

                if subject is None or len(subject.strip()) < 2:  # Что предмет есть
                    continue
                # print(subject)
                us.get_group_day_schedule(group, day, week_type).add_lesson(time, subject)
                col_prev_subject[group] = subject
        print(us)
        return us


if __name__ == '__main__':
    # pprint(parse_excel_file('./rasp.xlsx'))
    parser = Parser()
    grp_shedule: UniversitySchedule = parser.parse_file('./rasp.xlsx')
    print(grp_shedule)
